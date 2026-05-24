import openpyxl
import os


class ExcelReader:
    """Utility to read test data from Excel file."""

    def __init__(self, file_name="test_data.xlsx"):
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.file_path = os.path.join(base_dir, "test_data", file_name)
        self.wb = openpyxl.load_workbook(self.file_path)

    def get_sheet_data(self, sheet_name):
        """Return all rows from a sheet as list of dicts (header as keys)."""
        ws = self.wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() for h in rows[0]]
        result = []
        for row in rows[1:]:
            row_dict = dict(zip(headers, row))
            result.append(row_dict)
        return result

    def get_login_data(self):
        """Return all rows from LoginData sheet."""
        return self.get_sheet_data("LoginData")

    def get_filter_data(self):
        """Return all rows from FilterData sheet."""
        return self.get_sheet_data("FilterData")

    def get_test_cases(self):
        """Return all rows from TestCases sheet."""
        return self.get_sheet_data("TestCases")

    def write_result(self, tc_id, status, failure_reason="", execution_date="", executed_by=""):
        """Write execution result back to ExecutionResults sheet."""
        import datetime
        ws = self.wb["ExecutionResults"]
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == str(tc_id):
                row[2].value = execution_date or datetime.datetime.now().strftime("%d-%m-%Y %H:%M")
                row[3].value = executed_by or "Automation"
                row[4].value = status
                row[5].value = failure_reason
                break
        self.wb.save(self.file_path)
        print(f"[ExcelReader] Result written for {tc_id}: {status}")


if __name__ == "__main__":
    reader = ExcelReader()
    print("=== Test Cases ===")
    for tc in reader.get_test_cases():
        print(tc)
    print("\n=== Login Data ===")
    for ld in reader.get_login_data():
        print(ld)
